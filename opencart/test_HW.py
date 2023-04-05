import openpyxl
import pytest
import time
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from opencart.pageobject import loginpage
@pytest.fixture()
def setup():
    print("this run before every method")
def test_def1(setup):
    print("hello world")
def test_loginopencart1(setup):
        excelpath = "C:\selenium\pydemo.xlsx"
        workbook = openpyxl.load_workbook(excelpath)
        sheet = workbook.active
        rows = sheet.max_row
        cols = sheet.max_column
        print(rows)
        for i in range(2, rows + 1):
            print(i)
            uname = sheet.cell(row=i, column=1).value
            pword = sheet.cell(row=i, column=2).value
            driver=webdriver.Chrome(ChromeDriverManager().install())
            driver.get("https://demo.opencart.com/")
            lp = loginpage(driver)
            lp.clickonmyaccount()
            lp.clickonloginlink()
            time.sleep(10)
            lp.enteremail(uname)
            lp.enterpassword(pword)
            lp.clickonsubmit()
            time.sleep(5)
         # # driver.find_element(By.LINK_TEXT,"My Account").click()
         # # driver.find_element(By.LINK_TEXT,"Login").click()
         # pageobject.lp.clickonmyaccount(self)
         # pageobject.lp.clickonloginlink(self)
         # time.sleep(5)
         # if self.title=="Your Store":
         #     print("title not matched")
         # else:
         #     print("title matched")
         #
         # pageobject.lp.enteremail(self,uname)
         # pageobject.lp.enterpassword(self,pword)
         # pageobject.lp.clickonsubmit(self)
         # # driver.find_element(By.ID,"input-email").send_keys(username)
         # # driver.find_element(By.ID,"input-password").send_keys(password)
         # # driver.find_element(By.XPATH,"//button[@type='submit']").click()
         # # driver.quit()
         # # #time.sleep(5)
         # pageobject.lp.loginpage(driver, uname, pword)
