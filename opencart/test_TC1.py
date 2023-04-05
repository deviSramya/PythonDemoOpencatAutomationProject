import openpyxl
import pytest
import time
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from opencart.pageobject import loginpage
from opencart.pageobject2 import registerpage
@pytest.fixture()
def setup():
    print("this run before every method")
def test_def1(setup):
    print("hello world")
def test_registeropencart(setup):
            excelpath = "C:\selenium\pydemo2.xlsx"
            workbook=openpyxl.load_workbook(excelpath)
            sheet=workbook.active
            rows=sheet.max_row
            cols=sheet.max_column
            for i in range(2,rows+1):
                 print(i)
                 fname=sheet.cell(row=i,column=1).value
                 lname=sheet.cell(row=i,column=2).value
                 uname=sheet.cell(row=i,column=3).value
                 pword=sheet.cell(row=i,column=4).value

                 driver=webdriver.Chrome(ChromeDriverManager().install())
                 driver.get("https://demo.opencart.com/")
                 lp = registerpage(driver)
                 lp.clickonmyaccount()
                 lp.clickonregisterlink()
                 time.sleep(2)
                 lp.enterfname(fname)
                 lp.enterlname(lname)
                 lp.enteremail(uname)
                 lp.enterpassword(pword)
                 lp.clickyesorno()
                 lp.clickcheckbox()
                 lp.clickonsubmit()
                 time.sleep(2)

