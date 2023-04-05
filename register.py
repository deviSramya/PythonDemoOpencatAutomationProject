import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
def registerpage(nameone, nametwo, username, password):
    driver = webdriver.Chrome(ChromeDriverManager().install())
    driver.get("https://demo.opencart.com/")
    driver.find_element(By.LINK_TEXT, "My Account").click()
    driver.find_element(By.LINK_TEXT, "Register").click()
    #print(driver.find_element(By.LINK_TEXT, "My Account").get)
    if driver.title == "Your Store":
         print("title not matched")
    else:
         print("title matched")
    driver.find_element(By.ID, "input-firstname").send_keys(nameone)
    driver.find_element(By.ID, "input-lastname").send_keys(nametwo)
    driver.find_element(By.ID, "input-email").send_keys(username)
    driver.find_element(By.ID, "input-password").send_keys(password)
    driver.find_element(By.ID, "input-newsletter-yes").click()
    # #driver.find_element(By.XPATH, "//*[@id='form-register']/div/div/div/input").click()
    # driver.find_element(By.XPATH, "//button[@type='submit']").click()
    # driver.find_element((By.XPATH, "//*[@id='narbar-menu']/ul/li[1]/a")).click()
    # driver.find_element((By.XPATH, "//*[@id='narbar-menu']/ul/li[1]/div/div/ul/li[2]/a")).click()
    driver.find_element(By.XPATH, "//button[@type='submit']").click()
    time.sleep(5)
    driver.quit()



excelpath2="C:\selenium\pydemo2.xlsx"
workbook=openpyxl.load_workbook(excelpath2)
sheet=workbook.active
rows=sheet.max_row
cols=sheet.max_column
for i in range(2, rows+1):
    print(i)
    fname =sheet.cell(row=i, column=1).value
    lname =sheet.cell(row=i, column=2).value
    uname =sheet.cell(row=i, column=3).value
    pword =sheet.cell(row=i, column=4).value
    registerpage(fname, lname, uname, pword)
